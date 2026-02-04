"""
Report Generation

Generates accounting reports (CSV and Excel) from classified transactions.
Includes historical price enrichment and summary calculations.

Key Metrics:
- GROSS REVENUE: Sum of fiat values of INCOME transactions only
- TOTAL EXPENSES: Sum of fiat values of EXPENSE transactions only
- NET CASH FLOW: GROSS REVENUE - TOTAL EXPENSES

TRANSFER and UNKNOWN transactions are excluded from all totals.
"""

from __future__ import annotations

import csv
import io
import logging
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import httpx

logger = logging.getLogger(__name__)


# =============================================================================
# Data Structures
# =============================================================================

@dataclass
class EnrichedTransaction:
    """Transaction enriched with fiat values and classification."""

    # Original fields
    tx_hash: str
    block_number: int
    block_date: str
    block_time: str
    from_address: str
    to_address: str
    direction: str
    network: str
    asset: str
    amount_raw: str
    amount_decimal: str
    tx_fee_native: str
    tx_status: str
    category: str
    notes: str

    # Enriched fields
    usd_value: Optional[float] = None
    cad_value: Optional[float] = None
    gas_fee_usd: Optional[float] = None
    gas_fee_cad: Optional[float] = None
    transaction_type: str = "UNKNOWN"


# =============================================================================
# Price Oracle (CoinGecko)
# =============================================================================

class PriceOracle:
    """Fetches historical and current crypto prices from CoinGecko."""

    BASE_URL = "https://api.coingecko.com/api/v3"
    RATE_LIMIT_DELAY = 1.5  # Seconds between requests

    # Token ID mapping
    TOKEN_IDS = {
        "ETH": "ethereum", "BTC": "bitcoin", "USDC": "usd-coin",
        "USDT": "tether", "DAI": "dai", "WETH": "weth",
        "WBTC": "wrapped-bitcoin", "MATIC": "matic-network",
        "LINK": "chainlink", "UNI": "uniswap", "AAVE": "aave",
    }

    def __init__(self):
        self.client = httpx.Client(timeout=30.0)
        self._cache: dict[str, dict] = {}
        self._last_request = 0.0

    def _rate_limit(self):
        """Enforce rate limiting."""
        elapsed = time.time() - self._last_request
        if elapsed < self.RATE_LIMIT_DELAY:
            time.sleep(self.RATE_LIMIT_DELAY - elapsed)
        self._last_request = time.time()

    def get_historical_price(
        self,
        token: str,
        date: datetime,
        currencies: list[str] = None,
    ) -> dict[str, float]:
        """Get historical price for a token on a specific date."""
        currencies = currencies or ["usd", "cad"]
        token_id = self.TOKEN_IDS.get(token.upper())

        if not token_id:
            # Unknown token - treat as stablecoin
            return {c: 1.0 for c in currencies}

        # Check cache
        date_str = date.strftime("%d-%m-%Y")
        cache_key = f"{token_id}_{date_str}"
        if cache_key in self._cache:
            return self._cache[cache_key]

        # Fetch from API
        self._rate_limit()
        try:
            url = f"{self.BASE_URL}/coins/{token_id}/history"
            response = self.client.get(url, params={"date": date_str})
            response.raise_for_status()
            data = response.json()

            market_data = data.get("market_data", {})
            prices = market_data.get("current_price", {})

            result = {c: prices.get(c, 0) for c in currencies}
            self._cache[cache_key] = result
            return result

        except Exception as e:
            logger.warning(f"Failed to get price for {token}: {e}")
            return {c: 0 for c in currencies}

    def get_current_price(
        self,
        token: str,
        currencies: list[str] = None,
    ) -> dict[str, float]:
        """Get current price for a token."""
        currencies = currencies or ["usd", "cad"]
        token_id = self.TOKEN_IDS.get(token.upper())

        if not token_id:
            return {c: 1.0 for c in currencies}

        self._rate_limit()
        try:
            url = f"{self.BASE_URL}/simple/price"
            params = {"ids": token_id, "vs_currencies": ",".join(currencies)}
            response = self.client.get(url, params=params)
            response.raise_for_status()
            data = response.json()

            prices = data.get(token_id, {})
            return {c: prices.get(c, 0) for c in currencies}

        except Exception as e:
            logger.warning(f"Failed to get current price for {token}: {e}")
            return {c: 0 for c in currencies}

    def close(self):
        self.client.close()


# Singleton instance
_price_oracle: Optional[PriceOracle] = None


def get_price_oracle() -> PriceOracle:
    global _price_oracle
    if _price_oracle is None:
        _price_oracle = PriceOracle()
    return _price_oracle


# =============================================================================
# Totals Calculation
# =============================================================================

def calculate_totals(transactions: list[EnrichedTransaction]) -> dict:
    """
    Calculate accounting totals from classified transactions.

    Rules:
    - GROSS REVENUE = sum of fiat values of INCOME transactions only
    - TOTAL EXPENSES = sum of fiat values of EXPENSE transactions only
    - NET CASH FLOW = GROSS REVENUE - TOTAL EXPENSES
    - TRANSFER and UNKNOWN are excluded from all totals
    """
    gross_revenue_usd = 0.0
    gross_revenue_cad = 0.0
    total_expenses_usd = 0.0
    total_expenses_cad = 0.0
    total_gas_usd = 0.0
    total_gas_cad = 0.0

    for tx in transactions:
        if tx.transaction_type == "INCOME":
            gross_revenue_usd += tx.usd_value or 0
            gross_revenue_cad += tx.cad_value or 0
        elif tx.transaction_type == "EXPENSE":
            total_expenses_usd += tx.usd_value or 0
            total_expenses_cad += tx.cad_value or 0
            total_gas_usd += tx.gas_fee_usd or 0
            total_gas_cad += tx.gas_fee_cad or 0

    return {
        "gross_revenue_usd": round(gross_revenue_usd, 2),
        "gross_revenue_cad": round(gross_revenue_cad, 2),
        "total_expenses_usd": round(total_expenses_usd, 2),
        "total_expenses_cad": round(total_expenses_cad, 2),
        "net_cash_flow_usd": round(gross_revenue_usd - total_expenses_usd, 2),
        "net_cash_flow_cad": round(gross_revenue_cad - total_expenses_cad, 2),
        "total_gas_fees_usd": round(total_gas_usd, 2),
        "total_gas_fees_cad": round(total_gas_cad, 2),
        "transaction_count": len(transactions),
        "income_count": len([t for t in transactions if t.transaction_type == "INCOME"]),
        "expense_count": len([t for t in transactions if t.transaction_type == "EXPENSE"]),
        "transfer_count": len([t for t in transactions if t.transaction_type == "TRANSFER"]),
        "unknown_count": len([t for t in transactions if t.transaction_type == "UNKNOWN"]),
    }


# =============================================================================
# CSV Export
# =============================================================================

def export_to_csv(
    transactions: list[EnrichedTransaction],
    totals: dict,
    output_path: Optional[Path] = None,
) -> str:
    """
    Export transactions to CSV with summary.

    Returns CSV content as string. Optionally writes to file.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    direction_display = {"In": "INCOMING", "Out": "OUTGOING"}

    # Header
    writer.writerow([
        "Transaction Hash", "Block Number", "Date", "Time",
        "From Address", "To Address", "Direction", "Transaction Type",
        "Network", "Token", "Amount (Crypto)", "USD Value", "CAD Value",
        "Gas Fee (Native)", "Gas Fee (USD)", "Gas Fee (CAD)",
        "Status", "Category", "Notes",
    ])

    # Data rows
    for tx in transactions:
        writer.writerow([
            tx.tx_hash, tx.block_number, tx.block_date, tx.block_time,
            tx.from_address, tx.to_address,
            direction_display.get(tx.direction, tx.direction),
            tx.transaction_type, tx.network, tx.asset, tx.amount_decimal,
            f"{tx.usd_value:.2f}" if tx.usd_value else "",
            f"{tx.cad_value:.2f}" if tx.cad_value else "",
            tx.tx_fee_native,
            f"{tx.gas_fee_usd:.2f}" if tx.gas_fee_usd else "",
            f"{tx.gas_fee_cad:.2f}" if tx.gas_fee_cad else "",
            tx.tx_status, tx.category, tx.notes,
        ])

    # Summary section
    writer.writerow([])
    writer.writerow(["SUMMARY"])
    writer.writerow(["Total Transactions", totals["transaction_count"]])
    writer.writerow(["Income Transactions", totals["income_count"]])
    writer.writerow(["Expense Transactions", totals["expense_count"]])
    writer.writerow(["Transfer Transactions", totals["transfer_count"]])
    writer.writerow(["Unknown Transactions", totals["unknown_count"]])
    writer.writerow([])
    writer.writerow(["GROSS REVENUE (USD)", f"${totals['gross_revenue_usd']:.2f}"])
    writer.writerow(["GROSS REVENUE (CAD)", f"${totals['gross_revenue_cad']:.2f}"])
    writer.writerow([])
    writer.writerow(["TOTAL EXPENSES (USD)", f"${totals['total_expenses_usd']:.2f}"])
    writer.writerow(["TOTAL EXPENSES (CAD)", f"${totals['total_expenses_cad']:.2f}"])
    writer.writerow([])
    writer.writerow(["NET CASH FLOW (USD)", f"${totals['net_cash_flow_usd']:.2f}"])
    writer.writerow(["NET CASH FLOW (CAD)", f"${totals['net_cash_flow_cad']:.2f}"])
    writer.writerow([])
    writer.writerow(["Gas Fees (USD)", f"${totals['total_gas_fees_usd']:.2f}"])
    writer.writerow(["Gas Fees (CAD)", f"${totals['total_gas_fees_cad']:.2f}"])

    content = output.getvalue()

    if output_path:
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            f.write(content)

    return content


# =============================================================================
# Excel Export
# =============================================================================

def export_to_xlsx(
    transactions: list[EnrichedTransaction],
    totals: dict,
    output_path: Optional[Path] = None,
) -> bytes:
    """
    Export transactions to Excel with formatting.

    Returns XLSX bytes. Optionally writes to file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Styles
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    total_font = Font(bold=True)
    revenue_fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
    expense_fill = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")
    net_fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
    money_format = '#,##0.00'

    direction_display = {"In": "INCOMING", "Out": "OUTGOING"}

    # Headers
    headers = [
        "Transaction Hash", "Block #", "Date", "Time",
        "From Address", "To Address", "Direction", "Transaction Type",
        "Network", "Token", "Amount", "USD Value", "CAD Value",
        "Gas Fee", "Gas (USD)", "Gas (CAD)", "Status", "Category", "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, tx in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=tx.tx_hash)
        ws.cell(row=row_idx, column=2, value=tx.block_number)
        ws.cell(row=row_idx, column=3, value=tx.block_date)
        ws.cell(row=row_idx, column=4, value=tx.block_time)
        ws.cell(row=row_idx, column=5, value=tx.from_address)
        ws.cell(row=row_idx, column=6, value=tx.to_address)
        ws.cell(row=row_idx, column=7, value=direction_display.get(tx.direction, tx.direction))
        ws.cell(row=row_idx, column=8, value=tx.transaction_type)
        ws.cell(row=row_idx, column=9, value=tx.network)
        ws.cell(row=row_idx, column=10, value=tx.asset)
        ws.cell(row=row_idx, column=11, value=float(tx.amount_decimal) if tx.amount_decimal else 0)

        usd_cell = ws.cell(row=row_idx, column=12, value=tx.usd_value or 0)
        usd_cell.number_format = money_format

        cad_cell = ws.cell(row=row_idx, column=13, value=tx.cad_value or 0)
        cad_cell.number_format = money_format

        ws.cell(row=row_idx, column=14, value=float(tx.tx_fee_native) if tx.tx_fee_native else 0)

        gas_usd = ws.cell(row=row_idx, column=15, value=tx.gas_fee_usd or 0)
        gas_usd.number_format = money_format

        gas_cad = ws.cell(row=row_idx, column=16, value=tx.gas_fee_cad or 0)
        gas_cad.number_format = money_format

        ws.cell(row=row_idx, column=17, value=tx.tx_status)
        ws.cell(row=row_idx, column=18, value=tx.category)
        ws.cell(row=row_idx, column=19, value=tx.notes)

    # Summary section
    total_row = len(transactions) + 3

    ws.cell(row=total_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)

    ws.cell(row=total_row + 1, column=1, value="Total Transactions:")
    ws.cell(row=total_row + 1, column=2, value=totals["transaction_count"])

    ws.cell(row=total_row + 2, column=1, value="Income:")
    ws.cell(row=total_row + 2, column=2, value=totals["income_count"])

    ws.cell(row=total_row + 3, column=1, value="Expense:")
    ws.cell(row=total_row + 3, column=2, value=totals["expense_count"])

    ws.cell(row=total_row + 4, column=1, value="Transfer:")
    ws.cell(row=total_row + 4, column=2, value=totals["transfer_count"])

    ws.cell(row=total_row + 5, column=1, value="Unknown:")
    ws.cell(row=total_row + 5, column=2, value=totals["unknown_count"])

    # Gross Revenue
    rev_row = total_row + 7
    ws.cell(row=rev_row, column=1, value="GROSS REVENUE (USD)").font = total_font
    rev_usd = ws.cell(row=rev_row, column=2, value=totals["gross_revenue_usd"])
    rev_usd.font = total_font
    rev_usd.fill = revenue_fill
    rev_usd.number_format = '"$"#,##0.00'

    ws.cell(row=rev_row + 1, column=1, value="GROSS REVENUE (CAD)").font = total_font
    rev_cad = ws.cell(row=rev_row + 1, column=2, value=totals["gross_revenue_cad"])
    rev_cad.font = total_font
    rev_cad.fill = revenue_fill
    rev_cad.number_format = '"$"#,##0.00'

    # Total Expenses
    exp_row = rev_row + 3
    ws.cell(row=exp_row, column=1, value="TOTAL EXPENSES (USD)").font = total_font
    exp_usd = ws.cell(row=exp_row, column=2, value=totals["total_expenses_usd"])
    exp_usd.font = total_font
    exp_usd.fill = expense_fill
    exp_usd.number_format = '"$"#,##0.00'

    ws.cell(row=exp_row + 1, column=1, value="TOTAL EXPENSES (CAD)").font = total_font
    exp_cad = ws.cell(row=exp_row + 1, column=2, value=totals["total_expenses_cad"])
    exp_cad.font = total_font
    exp_cad.fill = expense_fill
    exp_cad.number_format = '"$"#,##0.00'

    # Net Cash Flow
    net_row = exp_row + 3
    ws.cell(row=net_row, column=1, value="NET CASH FLOW (USD)").font = total_font
    net_usd = ws.cell(row=net_row, column=2, value=totals["net_cash_flow_usd"])
    net_usd.font = total_font
    net_usd.fill = net_fill
    net_usd.number_format = '"$"#,##0.00'

    ws.cell(row=net_row + 1, column=1, value="NET CASH FLOW (CAD)").font = total_font
    net_cad = ws.cell(row=net_row + 1, column=2, value=totals["net_cash_flow_cad"])
    net_cad.font = total_font
    net_cad.fill = net_fill
    net_cad.number_format = '"$"#,##0.00'

    # Gas Fees
    gas_row = net_row + 3
    ws.cell(row=gas_row, column=1, value="Gas Fees (USD)")
    gas_usd_total = ws.cell(row=gas_row, column=2, value=totals["total_gas_fees_usd"])
    gas_usd_total.number_format = '"$"#,##0.00'

    ws.cell(row=gas_row + 1, column=1, value="Gas Fees (CAD)")
    gas_cad_total = ws.cell(row=gas_row + 1, column=2, value=totals["total_gas_fees_cad"])
    gas_cad_total.number_format = '"$"#,##0.00'

    # Column widths
    column_widths = {
        1: 20, 2: 10, 3: 12, 4: 10, 5: 20, 6: 20, 7: 12, 8: 16, 9: 10, 10: 8,
        11: 15, 12: 12, 13: 12, 14: 12, 15: 12, 16: 12, 17: 10, 18: 15, 19: 30,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(xlsx_bytes)

    return xlsx_bytes


# =============================================================================
# Export Service (Orchestrator)
# =============================================================================

class ExportService:
    """
    Orchestrates the complete export workflow.

    Usage:
        service = ExportService(output_dir="./exports")
        result = service.generate_export(
            wallet_addresses=["0x..."],
            blockchain="ethereum",
            api_key="...",
        )
    """

    def __init__(self, output_dir: str | Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.price_oracle = get_price_oracle()

    def enrich_transactions(
        self,
        transactions: list,
        include_prices: bool = True,
    ) -> list[EnrichedTransaction]:
        """Enrich transactions with fiat values."""
        enriched = []

        for tx in transactions:
            enriched_tx = EnrichedTransaction(
                tx_hash=tx.tx_hash,
                block_number=tx.block_number,
                block_date=tx.block_date,
                block_time=tx.block_time,
                from_address=tx.from_address,
                to_address=tx.to_address,
                direction=tx.direction,
                network=tx.network,
                asset=tx.asset,
                amount_raw=tx.amount_raw,
                amount_decimal=tx.amount_decimal,
                tx_fee_native=tx.tx_fee_native,
                tx_status=tx.tx_status,
                category=tx.category,
                notes=tx.notes,
            )

            if include_prices and tx.block_date:
                try:
                    tx_date = datetime.strptime(tx.block_date, "%Y-%m-%d")
                    prices = self.price_oracle.get_historical_price(
                        tx.asset, tx_date, ["usd", "cad"]
                    )

                    amount = float(tx.amount_decimal) if tx.amount_decimal else 0
                    enriched_tx.usd_value = amount * prices.get("usd", 0)
                    enriched_tx.cad_value = amount * prices.get("cad", 0)

                    # Gas fees
                    if tx.tx_fee_native and tx.direction == "Out":
                        fee = float(tx.tx_fee_native)
                        native = "ETH" if tx.network == "Ethereum" else "BTC"
                        native_prices = self.price_oracle.get_historical_price(
                            native, tx_date, ["usd", "cad"]
                        )
                        enriched_tx.gas_fee_usd = fee * native_prices.get("usd", 0)
                        enriched_tx.gas_fee_cad = fee * native_prices.get("cad", 0)

                except Exception as e:
                    logger.warning(f"Failed to enrich {tx.tx_hash[:16]}: {e}")

            enriched.append(enriched_tx)

        return enriched

    def classify_transactions(
        self,
        transactions: list[EnrichedTransaction],
        wallet_addresses: list[str],
        overrides: Optional[dict[str, str]] = None,
    ) -> None:
        """Classify transactions in place."""
        wallet_set = {w.lower() for w in wallet_addresses}
        overrides = overrides or {}

        for tx in transactions:
            if tx.tx_hash in overrides:
                override = overrides[tx.tx_hash].upper()
                if override in ("INCOME", "EXPENSE", "TRANSFER", "UNKNOWN"):
                    tx.transaction_type = override
                    continue

            from_lower = tx.from_address.lower() if tx.from_address else ""
            to_lower = tx.to_address.lower() if tx.to_address else ""

            if from_lower in wallet_set and to_lower in wallet_set:
                tx.transaction_type = "TRANSFER"
            elif tx.direction == "In":
                tx.transaction_type = "INCOME"
            elif tx.direction == "Out":
                tx.transaction_type = "EXPENSE"
            else:
                tx.transaction_type = "UNKNOWN"

    def generate_export(
        self,
        wallet_addresses: list[str],
        blockchain: str,
        api_key: Optional[str] = None,
        from_date: Optional[datetime] = None,
        to_date: Optional[datetime] = None,
        output_format: str = "csv",
        include_prices: bool = True,
        classification_overrides: Optional[dict[str, str]] = None,
        classification_addresses: Optional[list[str]] = None,
    ) -> dict:
        """
        Complete export workflow.

        Args:
            wallet_addresses: Addresses to fetch transactions for
            blockchain: Network (ethereum, bitcoin)
            api_key: Optional API key for the blockchain provider
            from_date: Optional start date filter
            to_date: Optional end date filter
            output_format: "csv" or "xlsx"
            include_prices: Whether to fetch historical prices
            classification_overrides: Manual tx_hash -> type overrides
            classification_addresses: All user wallet addresses for transfer detection.
                                      If provided, used instead of wallet_addresses for
                                      determining if a tx is a TRANSFER between own wallets.

        Returns dict with filename, totals, and transaction count.
        """
        from src.wallet import LedgerGenerator

        # Fetch transactions
        logger.info(f"Fetching transactions for {len(wallet_addresses)} wallet(s)...")
        all_rows = []

        for address in wallet_addresses:
            generator = LedgerGenerator(network=blockchain, api_key=api_key)
            rows = generator.generate_ledger(address, from_date, to_date)
            all_rows.extend(rows)
            generator.close()

        if not all_rows:
            return {"success": True, "message": "No transactions found", "transaction_count": 0}

        # Sort by date
        all_rows.sort(key=lambda r: (r.block_date, r.block_time))

        # Enrich with prices
        logger.info(f"Enriching {len(all_rows)} transactions...")
        enriched = self.enrich_transactions(all_rows, include_prices=include_prices)

        # Classify - use classification_addresses if provided (includes all user's wallets)
        # This enables automatic transfer detection between any of the user's wallets
        addresses_for_classification = classification_addresses or wallet_addresses
        logger.info(f"Classifying transactions (using {len(addresses_for_classification)} addresses for transfer detection)...")
        self.classify_transactions(enriched, addresses_for_classification, classification_overrides)

        # Calculate totals
        totals = calculate_totals(enriched)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"chainledger_export_{timestamp}.{output_format}"
        filepath = self.output_dir / filename

        # Export
        if output_format.lower() == "xlsx":
            export_to_xlsx(enriched, totals, filepath)
        else:
            export_to_csv(enriched, totals, filepath)

        logger.info(f"Export complete: {filename}")

        return {
            "success": True,
            "filename": filename,
            "filepath": str(filepath),
            "transaction_count": len(enriched),
            "totals": totals,
        }
